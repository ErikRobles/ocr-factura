"""
Export extracted rows to .xlsx with sheets: Extracted, Needs_Review, Run_Summary.
Uses openpyxl; adds clickable hyperlinks to original image paths.
"""
from pathlib import Path
from typing import List
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from .models import ExtractedRow

# Column order: high-priority first (per spec)
COLUMNS = [
    "Retailer_Name",
    "Ticket_or_Folio",
    "Ticket_or_Folio_Candidates",
    "Amount_Total",
    "Amount_Subtotal",
    "Currency",
    "Payment_Method",
    "Date",
    "Time",
    "Transaction_Number",
    "Sucursal",
    "Tienda",
    "Caja",
    "Cajero",
    "Facturacion_URL_On_Ticket",
    "Facturacion_Email",
    "Facturacion_Method",
    "Facturacion_Notes",
    "OCR_Quality",
    "Confidence_Score",
    "Missing_Fields",
    "Warnings",
    "Image_File_Name",
    "Image_File_Path",
    "Image_Open_Link",
]


# Map column name -> ExtractedRow attribute (snake_case)
def _row_to_dict(row: ExtractedRow) -> dict:
    return {
        "Retailer_Name": row.retailer_name,
        "Ticket_or_Folio": row.ticket_or_folio,
        "Ticket_or_Folio_Candidates": row.ticket_or_folio_candidates,
        "Amount_Total": row.amount_total,
        "Amount_Subtotal": row.amount_subtotal,
        "Currency": row.currency,
        "Payment_Method": row.payment_method,
        "Date": row.date,
        "Time": row.time,
        "Transaction_Number": row.transaction_number,

        # Spanish-friendly naming
        "Sucursal": getattr(row, "sucursal", ""),
        "Tienda": getattr(row, "tienda", ""),
        "Caja": getattr(row, "caja", ""),
        "Cajero": getattr(row, "cajero", ""),

        "Facturacion_URL_On_Ticket": row.facturacion_url_on_ticket,
        "Facturacion_Email": getattr(row, "email_on_ticket", ""),
        "Facturacion_Method": row.facturacion_method,
        "Facturacion_Notes": row.facturacion_notes,
        "OCR_Quality": row.ocr_quality,
        "Confidence_Score": row.confidence_score,
        "Missing_Fields": row.missing_fields,
        "Warnings": row.warnings,
        "Image_File_Name": row.image_file_name,
        "Image_File_Path": row.image_file_path,
        "Image_Open_Link": row.image_open_link,
    }



def _write_sheet_rows(ws, rows: List[ExtractedRow], write_header: bool = True):
    """Write header and data rows; add hyperlink in Image_Open_Link column."""
    if write_header:
        for c, col in enumerate(COLUMNS, 1):
            ws.cell(row=1, column=c, value=col)
        start_row = 2
    else:
        start_row = 1
    link_col_idx = COLUMNS.index("Image_Open_Link") + 1
    for r, row in enumerate(rows, start_row):
        d = _row_to_dict(row)
        for c, col in enumerate(COLUMNS, 1):
            val = d.get(col, "")
            cell = ws.cell(row=r, column=c, value=val)
            if col == "Image_Open_Link" and val and Path(val).exists():
                cell.hyperlink = val
                cell.font = Font(underline="single", color="0563C1")
    return start_row + len(rows) - 1


def export_to_excel(
    rows: List[ExtractedRow],
    output_path: Path,
    run_warnings: List[str],
    run_failures: int = 0,
) -> Path:
    """
    Write workbook with sheets Extracted, Needs_Review, Run_Summary.
    output_path: path for .xlsx file.
    Returns output_path.
    """
    wb = Workbook()
    # Extracted (all rows)
    ws_all = wb.active
    ws_all.title = "Extracted"
    _write_sheet_rows(ws_all, rows, write_header=True)

    # Needs_Review (subset)
    need_review = [r for r in rows if r.is_low_confidence_or_missing_key_fields()]
    ws_review = wb.create_sheet("Needs_Review")
    _write_sheet_rows(ws_review, need_review, write_header=True)

    # Run_Summary
    ws_sum = wb.create_sheet("Run_Summary")
    ws_sum["A1"] = "Metric"
    ws_sum["B1"] = "Value"
    summary = [
        ("Total images processed", len(rows)),
        ("Needs review count", len(need_review)),
        ("Failures / unreadable", run_failures),
        ("Run timestamp", datetime.now().isoformat()),
    ]
    for i, (label, value) in enumerate(summary, 2):
        ws_sum.cell(row=i, column=1, value=label)
        ws_sum.cell(row=i, column=2, value=value)
    if run_warnings:
        start = len(summary) + 2
        ws_sum.cell(row=start, column=1, value="Warnings")
        for i, w in enumerate(run_warnings, start + 1):
            ws_sum.cell(row=i, column=1, value=w)

    wb.save(str(output_path))
    return output_path


def default_output_name(folder: Path) -> str:
    """Default filename: MonthlyInvoicing_YYYY-MM_<timestamp>.xlsx"""
    now = datetime.now()
    return f"MonthlyInvoicing_{now.year}-{now.month:02d}_{now.strftime('%H%M%S')}.xlsx"
