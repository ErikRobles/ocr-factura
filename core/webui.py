"""
Tiny local Web UI (Option B) for:
- paste ChatGPT JSON batch
- validate + add to session
- merge all batches
- export one Excel

No APIs. No scraping. User manually posts images in ChatGPT and pastes JSON here.
"""

from __future__ import annotations

import json
import os
import webbrowser
from dataclasses import asdict
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Tuple

import html as html_module
from flask import Flask, request, jsonify, Response
from openpyxl import load_workbook

from .json_cleaner import repair_json_text
from .chatgpt_batch_schema import validate_batch  # should exist in your repo
from .merge_batches import merge_batches          # should exist in your repo
from .excel_export import export_to_excel, default_output_name
from .models import ExtractedRow


# Prompt to show in UI so users can copy it into ChatGPT before uploading receipt images
CHATGPT_HINT = r"""You are an AI extraction engine designed for a private Mexican receipt-to-invoice workflow.

This is NOT a chat task.
This is NOT a summary task.
This is NOT OCR transcription.

You MUST extract ONLY invoicing-relevant data and return STRICT VALID JSON for batch processing.

The output will be pasted directly into a local tool that validates JSON and merges into Excel.
If JSON is malformed, the system breaks.

Follow these instructions EXACTLY.

==================================================
CORE OBJECTIVE

From 3–5 uploaded Mexican receipt images:

Extract ONLY the data required to generate invoices (facturas).
Return ONE clean JSON object.
No explanations.
No markdown.
No extra text.

If a value cannot be found → use "" (empty string).
Never invent values.
Never guess.

==================================================
OUTPUT STRUCTURE (MANDATORY)

Return ONLY this JSON structure:

{
"batch_meta": {
"batch_id": "",
"image_count": 0
},
"rows": []
}

Each row inside "rows" must be:

{
"image_file_name": "",
"retailer_name": "",
"ticket_number": "",
"date": "",
"time": "",
"amount_total": "",
"amount_subtotal": "",
"transaction_number": "",
"payment_method": "",
"cashier": "",
"caja": "",
"tienda": "",
"sucursal": "",
"complejo": "",
"facturacion_url": "",
"facturacion_email": "",
"warnings": ""
}

==================================================
GLOBAL EXTRACTION RULES

OUTPUT RULES

Output ONLY valid JSON

No markdown

No comments

No explanations

No trailing commas

Must parse with JSON.parse()

LANGUAGE
Receipts are usually in Spanish.
Extract values exactly as printed.

DATE FORMAT
Convert to:
YYYY-MM-DD

Examples:
02/02/2026 → 2026-02-02
5/1/2026 → 2026-05-01 if clearly May 1
If ambiguous → copy exactly as printed

TIME FORMAT
Keep exactly as printed (HH:MM or HH:MM:SS)

AMOUNTS
Use numbers exactly as printed.
No currency symbols.
No commas unless printed.

Example:
197.00
482.95

DO NOT CONFUSE

RFC (never ticket number)

Authorization codes

Card numbers

Store phone numbers

Order numbers unrelated to invoicing

USE EMPTY STRING "" IF:

Value not present

Unclear

Obstructed

Multiple conflicting values

Never guess.

==================================================
RETAILER NAME NORMALIZATION

Use public brand names ONLY:

CHEDRAUI
CINEMEX
KIKOS PASTES
WALMART
OXXO
ITALCAFE
H&M
etc.

Remove:
S.A. de C.V.
legal entity names
long fiscal names

If Cinemex logo not obvious:
If website shows cinemex.com → retailer_name = CINEMEX

==================================================
RETAILER-SPECIFIC EXTRACTION RULES
========================
CINEMEX

Cinemex requires:

Ticket number

Date

Complejo

Total amount

Where to find:

Ticket number:
Usually near bottom labeled:
"Ticket:"
Example: Ticket:1190528

Complejo:
Often shown as:
"Complejo: OCE"
or near ticket number
Extract exactly.

Date & time:
Near ticket line:
Example:
05/01/2026 19:51:30

Monto:
Use TOTAL paid or Amount Due

Retailer identification:
Cinemex name may be small.
If receipt shows:
cinemex.com
→ retailer_name = CINEMEX

==================================================
CHEDRAUI

Critical field: FOLIO (ticket_number)

Chedraui requires:
Ticket number = FOLIO (20 digits)

Rules:

Located near bottom

Above barcode

Preceded by "FOLIO"

Often long numeric string

Example:
FOLIO: 2602 0212 1801 8206 0081

Remove spaces.
Keep full number.

DO NOT use:
Authorization
Card number
Transaction number

Monto:
Use TOTAL M.N.

Date/time:
Near bottom line with register info.

==================================================
KIKOS PASTES

No online invoice portal.
Invoice requested by EMAIL.

Extract:
facturacion_email

Look for:
"Enviar una fotografía..."
"facturacion@..."
Extract exact email.

ticket_number:
Use:
VENTA
Folio
Ticket
or Venta number

If none clear:
Use transaction/venta number.

==================================================
COMMON FIELD DETECTION

ticket_number may appear as:
FOLIO
TICKET
VENTA
OPER
TRANSACCION
NOTA
ORDEN

transaction_number:
Only if clearly labeled:
TR:
TRANS
OPER
AUTH
APROBACION

caja:
Extract if labeled CAJA

cashier:
Look for:
CAJERO
LE ATENDIO
ATENDIO

payment_method:
TARJETA
EFECTIVO
DEBITO
CREDITO

amount_total:
Use:
TOTAL
IMPORTE TOTAL
TOTAL M.N.
AMOUNT DUE

amount_subtotal:
Use only if clearly labeled SUBTOTAL

==================================================
FACTURACION URL

Extract if present:
facturacion portal URL
website for invoices

Examples:
chedraui.com.mx
cinemex.com
facturacion....

If email-only → leave URL empty and fill email.

==================================================
WARNINGS FIELD

Use ONLY if something unusual:

Examples:
"ticket number unclear"
"folio partially cut"
"multiple totals found"
"cinemex inferred from website"

Otherwise:
"" (empty)

==================================================
CRITICAL JSON RULES

Return ONE JSON object only.
No extra output.
No explanations.

batch_meta.image_count must equal number of images processed.
One row per image.

Everything must be valid JSON.
Single final answer must parse cleanly.

If you break JSON formatting → system fails.

Proceed carefully."""

# Where we persist a running session of pasted batches (exe dir when frozen)
from .paths import get_base_dir

_REPO_ROOT = get_base_dir()
_SESSION_DIR = _REPO_ROOT / "output" / "sessions"
_OUTPUT_DIR = _REPO_ROOT / "output"
_SESSION_DIR.mkdir(parents=True, exist_ok=True)


def _list_excel_files() -> List[Dict[str, str]]:
    """List .xlsx files in output dir, newest first. Returns [{"name": basename, "path": str}, ...]."""
    if not _OUTPUT_DIR.is_dir():
        return []
    files = []
    for p in _OUTPUT_DIR.glob("*.xlsx"):
        if p.is_file():
            try:
                mtime = p.stat().st_mtime
                files.append({"name": p.name, "path": str(p.resolve()), "mtime": mtime})
            except OSError:
                pass
    files.sort(key=lambda x: x["mtime"], reverse=True)
    return [{"name": f["name"], "path": f["path"]} for f in files]


def _read_excel_extracted(path: Path) -> Tuple[List[str], List[Dict[str, Any]], Dict[str, Any]]:
    """Read Excel at path; return (headers, rows, summary_dict). Raises if sheet missing."""
    wb = load_workbook(path, read_only=True, data_only=True)
    if "Extracted" not in wb.sheetnames:
        wb.close()
        raise ValueError("No 'Extracted' sheet")
    ws = wb["Extracted"]
    rows_iter = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows_iter:
        return [], [], {}
    headers = [str(h) if h is not None else "" for h in rows_iter[0]]
    rows = []
    for r in rows_iter[1:]:
        row = {}
        for i, h in enumerate(headers):
            val = r[i] if i < len(r) else None
            row[h] = "" if val is None else str(val).strip()
        rows.append(row)
    summary = {"total_rows": len(rows), "file_name": path.name}
    return headers, rows, summary


def _session_file() -> Path:
    # One session per day by default (simple + safe)
    day = datetime.now().strftime("%Y-%m-%d")
    return _SESSION_DIR / f"session_{day}.jsonl"


def _append_session_line(obj: Dict[str, Any]) -> None:
    p = _session_file()
    with p.open("a", encoding="utf-8") as f:
        f.write(json.dumps(obj, ensure_ascii=False) + "\n")


def _load_session_batches() -> List[List[Dict[str, Any]]]:
    p = _session_file()
    if not p.exists():
        return []
    batches: List[List[Dict[str, Any]]] = []
    with p.open("r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            rec = json.loads(line)
            if rec.get("type") == "batch" and isinstance(rec.get("items"), list):
                batches.append(rec["items"])
    return batches


def _to_rows(merged_items: List[Dict[str, Any]]) -> List[ExtractedRow]:
    """
    Convert merged JSON dicts (ChatGPT batch schema) -> ExtractedRow objects.
    Keeps only fields your Excel exporter expects.
    """
    rows: List[ExtractedRow] = []
    for item in merged_items:
        r = ExtractedRow()

        # Required-ish
        r.image_file_name = item.get("image_file_name") or ""
        r.retailer_name = item.get("retailer_name") or "Unable to read value"
        r.ticket_or_folio = item.get("ticket_number") or item.get("ticket_or_folio") or "Unable to read value"


        # Amounts (keep as strings in row model if your exporter expects strings)
        amt_total = item.get("amount_total")
        r.amount_total = "Unable to read value" if amt_total in (None, "") else str(amt_total)

        amt_sub = item.get("amount_subtotal")
        r.amount_subtotal = "" if amt_sub in (None, "") else str(amt_sub)

        # Optional
        r.date = item.get("date") or ""
        r.time = item.get("time") or ""
        r.transaction_number = item.get("transaction_number") or ""
        r.payment_method = item.get("payment_method") or ""
        r.currency = item.get("currency") or ""

        r.store_branch = item.get("sucursal") or item.get("store_branch") or ""
        r.cashier_number = item.get("cashier") or item.get("cajero") or item.get("cashier_number") or ""

        r.register_number = item.get("caja") or item.get("register_number") or ""

        # Facturación (URL and/or email, e.g. Kikos: "Enviar fotografía al correo facturacion@...")
        r.facturacion_url_on_ticket = item.get("facturacion_url") or ""
        r.email_on_ticket = item.get("facturacion_email") or item.get("email_on_ticket") or ""
        r.facturacion_notes = item.get("facturacion_instructions") or item.get("notes") or ""
        r.facturacion_method = "URL" if r.facturacion_url_on_ticket else ("EMAIL" if r.email_on_ticket else "UNKNOWN")

        # Quality / review
        conf = item.get("confidence")
        try:
            r.confidence_score = float(conf) if conf is not None else 0.5
        except Exception:
            r.confidence_score = 0.5

        needs_review = bool(item.get("needs_review")) or r.confidence_score < 0.70
        # Keep your model's logic too (it may check missing fields)
        r.ocr_quality = "Good" if r.confidence_score >= 0.80 else ("Medium" if r.confidence_score >= 0.50 else "Poor")
        r.warnings = "; ".join(item.get("warnings") or []) if isinstance(item.get("warnings"), list) else (item.get("warnings") or "")

        # Links (optional)
        img_path = item.get("image_file_path") or ""
        r.image_file_path = img_path
        r.image_open_link = img_path

        # Build missing fields string for Needs_Review
        missing = []
        if not r.retailer_name or r.retailer_name == "Unable to read value":
            missing.append("Retailer_Name")
        if not r.ticket_or_folio or r.ticket_or_folio == "Unable to read value":
            missing.append("Ticket_or_Folio")
        if not r.amount_total or r.amount_total == "Unable to read value":
            missing.append("Amount_Total")
        r.missing_fields = ",".join(missing)

        # If needs_review but not reflected by missing_fields logic, add warning
        if needs_review and "needs_review" not in (r.warnings or "").lower():
            pass

        rows.append(r)

    # Sort predictable
    rows.sort(key=lambda x: (x.retailer_name or "", x.date or "", x.image_file_name or ""))
    return rows


def _view_dashboard_html(file_name: str) -> str:
    """Return HTML for the visualize dashboard; fetches data for file_name via API."""
    file_name_js = json.dumps(file_name)  # safe for JS string
    return f"""<!DOCTYPE html>
<html lang="es">
<head>
  <meta charset="utf-8" />
  <title>OCRFactura — Vista de datos</title>
  <style>
    :root {{ --bg: #0f1419; --card: #1a2332; --border: #2d3a4d; --text: #e6edf3; --muted: #8b949e; --accent: #58a6ff; --green: #3fb950; --amber: #d29922; }}
    * {{ box-sizing: border-box; }}
    body {{ font-family: 'Segoe UI', system-ui, -apple-system, sans-serif; margin: 0; background: var(--bg); color: var(--text); min-height: 100vh; }}
    .container {{ max-width: 1200px; margin: 0 auto; padding: 24px; }}
    h1 {{ font-size: 1.75rem; font-weight: 600; margin: 0 0 8px 0; }}
    .sub {{ color: var(--muted); font-size: 0.95rem; margin-bottom: 24px; }}
    .summary {{ display: flex; gap: 16px; flex-wrap: wrap; margin-bottom: 24px; }}
    .summary-card {{ background: var(--card); border: 1px solid var(--border); border-radius: 12px; padding: 16px 20px; min-width: 140px; }}
    .summary-card .val {{ font-size: 1.5rem; font-weight: 700; color: var(--accent); }}
    .summary-card .label {{ font-size: 0.8rem; color: var(--muted); text-transform: uppercase; letter-spacing: 0.05em; margin-top: 4px; }}
    .retailer-group {{ background: var(--card); border: 1px solid var(--border); border-radius: 12px; margin-bottom: 16px; overflow: hidden; }}
    .retailer-group h2 {{ margin: 0; padding: 12px 20px; font-size: 1.1rem; background: rgba(88,166,255,0.08); border-bottom: 1px solid var(--border); }}
    .table-wrap {{ overflow-x: auto; }}
    table {{ width: 100%; border-collapse: collapse; font-size: 0.9rem; }}
    th {{ text-align: left; padding: 10px 16px; color: var(--muted); font-weight: 600; border-bottom: 1px solid var(--border); white-space: nowrap; }}
    td {{ padding: 10px 16px; border-bottom: 1px solid var(--border); }}
    tr:hover {{ background: rgba(88,166,255,0.04); }}
    .amt {{ font-variant-numeric: tabular-nums; color: var(--green); }}
    .loading {{ padding: 48px; text-align: center; color: var(--muted); }}
    .err {{ color: #f85149; padding: 24px; }}
    a {{ color: var(--accent); text-decoration: none; }}
    a:hover {{ text-decoration: underline; }}
  </style>
</head>
<body>
  <div class="container">
    <h1>Facturación — Vista de datos</h1>
    <p class="sub" id="file-label">Cargando…</p>
    <div id="summary" class="summary"></div>
    <div id="content"></div>
  </div>
  <script>
    const file = {file_name_js};
    document.getElementById('file-label').textContent = 'Archivo: ' + file;

    fetch('/api/excel-data?file=' + encodeURIComponent(file))
      .then(r => r.ok ? r.json() : Promise.reject(r.statusText))
      .then(data => {{
        const {{ headers, rows, summary }} = data;
        const total = rows.length;
        let totalAmount = 0;
        rows.forEach(r => {{
          const a = parseFloat(r['Amount_Total'] || r['amount_total'] || '0');
          if (!isNaN(a)) totalAmount += a;
        }});

        document.getElementById('summary').innerHTML = `
          <div class="summary-card"><span class="val">${{total}}</span><span class="label">Recepciones</span></div>
          <div class="summary-card"><span class="val">${{totalAmount.toFixed(2)}}</span><span class="label">Total (MXN)</span></div>
        `;

        const byRetailer = {{}};
        rows.forEach(row => {{
          const name = row['Retailer_Name'] || row['retailer_name'] || 'Sin tienda';
          if (!byRetailer[name]) byRetailer[name] = [];
          byRetailer[name].push(row);
        }});

        const displayHeaders = ['Retailer_Name','Ticket_or_Folio','Date','Time','Amount_Total','Amount_Subtotal','Payment_Method','Facturacion_Email','Facturacion_URL_On_Ticket'];
        const h = headers && headers.length ? headers : displayHeaders;
        const key = (row, col) => row[col] !== undefined ? row[col] : row[col.replace(/_/g, ' ')] || '';
        const esc = s => String(s || '').replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;').replace(/"/g,'&quot;');

        let html = '';
        Object.keys(byRetailer).sort().forEach(retailer => {{
          const groupRows = byRetailer[retailer];
          html += `<div class="retailer-group"><h2>${{esc(retailer)}}</h2><div class="table-wrap"><table><thead><tr>${{h.map(c => `<th>${{esc(c.replace(/_/g, ' '))}}</th>`).join('')}}</tr></thead><tbody>`;
          groupRows.forEach(row => {{
            html += '<tr>' + h.map(col => {{
              let v = key(row, col);
              if (col === 'Amount_Total' || col === 'Amount_Subtotal') v = v ? '<span class="amt">' + esc(v) + '</span>' : '';
              else if (col === 'Facturacion_URL_On_Ticket' && v) v = '<a href="' + esc(v) + '" target="_blank" rel="noopener">' + esc(v) + '</a>';
              else if (col === 'Facturacion_Email' && v) v = '<a href="mailto:' + esc(v) + '">' + esc(v) + '</a>';
              else v = esc(v);
              return '<td>' + v + '</td>';
            }}).join('') + '</tr>';
          }});
          html += '</tbody></table></div></div>';
        }});
        document.getElementById('content').innerHTML = html || '<p class="loading">No hay filas en la hoja Extracted.</p>';
      }})
      .catch(err => {{
        document.getElementById('content').innerHTML = '<p class="err">Error al cargar datos: ' + (err || '') + '</p>';
      }});
  </script>
</body>
</html>"""


def run_webui(port: int = 5173) -> None:
    app = Flask(__name__)

    @app.get("/")
    def home() -> Response:
        html = """
<!doctype html>
<html lang="es">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <title>OCRFactura — ChatGPT Batch Merger</title>
  <style>
    :root { --bg: #0f1419; --card: #1a2332; --border: #2d3a4d; --text: #e6edf3; --muted: #8b949e; --accent: #58a6ff; --green: #3fb950; --red: #f85149; }
    * { box-sizing: border-box; }
    body { font-family: 'Segoe UI', system-ui, -apple-system, sans-serif; margin: 0; background: var(--bg); color: var(--text); min-height: 100vh; -webkit-font-smoothing: antialiased; }
    .container { max-width: 720px; margin: 0 auto; padding: clamp(16px, 4vw, 32px); }
    h1 { font-size: clamp(1.35rem, 4vw, 1.75rem); font-weight: 600; margin: 0 0 8px 0; }
    .sub { color: var(--muted); font-size: 0.95rem; line-height: 1.5; margin-bottom: 24px; }
    .card { background: var(--card); border: 1px solid var(--border); border-radius: 12px; padding: clamp(16px, 4vw, 24px); margin-bottom: 20px; }
    .card h2 { margin: 0 0 8px 0; font-size: 1.05rem; font-weight: 600; color: var(--text); }
    .card p { margin: 0 0 12px 0; color: var(--muted); font-size: 0.9rem; }
    textarea { width: 100%; min-height: 200px; padding: 12px 14px; font-family: ui-monospace, 'SF Mono', Menlo, Consolas, monospace; font-size: 13px; background: var(--bg); border: 1px solid var(--border); border-radius: 8px; color: var(--text); resize: vertical; }
    textarea::placeholder { color: var(--muted); }
    .hint-box { min-height: 280px; font-size: 12px; }
    button { padding: 12px 18px; margin: 0 8px 8px 0; font-size: 0.95rem; font-family: inherit; font-weight: 500; color: var(--bg); background: var(--accent); border: none; border-radius: 8px; cursor: pointer; -webkit-tap-highlight-color: transparent; }
    button:hover { filter: brightness(1.1); }
    button:active { filter: brightness(0.95); }
    button.secondary { background: transparent; color: var(--accent); border: 1px solid var(--border); }
    button.secondary:hover { background: rgba(88,166,255,0.1); }
    #btn-visualize:disabled { opacity: 0.5; cursor: not-allowed; filter: none; }
    .row { margin: 12px 0; }
    .btn-row { display: flex; flex-wrap: wrap; gap: 8px; align-items: center; margin-top: 16px; }
    .ok { color: var(--green); font-weight: 500; }
    .bad { color: var(--red); font-weight: 500; }
    #status { min-height: 1.5em; }
    pre { background: var(--card); border: 1px solid var(--border); border-radius: 8px; padding: 14px; overflow: auto; font-size: 12px; color: var(--muted); margin: 0; white-space: pre-wrap; word-break: break-all; }
    .copy-feedback { margin-left: 8px; color: var(--green); font-size: 0.9rem; }
    .modal { position: fixed; inset: 0; z-index: 100; display: flex; align-items: center; justify-content: center; padding: 16px; }
    .modal-backdrop { position: absolute; inset: 0; background: rgba(0,0,0,0.6); }
    .modal-box { position: relative; background: var(--card); border: 1px solid var(--border); padding: 24px; border-radius: 12px; box-shadow: 0 16px 48px rgba(0,0,0,0.4); min-width: 280px; max-width: 100%; }
    .modal-box h3 { margin: 0 0 16px 0; font-size: 1.1rem; color: var(--text); }
    .modal-select { width: 100%; padding: 12px 14px; margin-bottom: 16px; font-size: 14px; background: var(--bg); border: 1px solid var(--border); border-radius: 8px; color: var(--text); }
    .modal-actions { display: flex; gap: 10px; justify-content: flex-end; flex-wrap: wrap; }
  </style>
</head>
<body>
  <div class="container">
    <h1>OCRFactura — Paste ChatGPT JSON</h1>
    <p class="sub">1) Copy the prompt below into ChatGPT, then post 3–5 receipt images.<br/>
       2) Copy the JSON response from ChatGPT.<br/>
       3) Paste it below and click <strong>Add Batch</strong>.<br/>
       4) Repeat. Then <strong>Export Excel</strong>.</p>

    <div class="card">
      <h2>Prompt for ChatGPT</h2>
      <p>Copy this into ChatGPT <em>before</em> sending your receipt images (3–5 per batch).</p>
      <textarea id="chatgpt-hint" readonly rows="28" class="hint-box">___CHATGPT_HINT_ESCAPED___</textarea>
      <div class="btn-row">
        <button type="button" onclick="copyChatGPTHint()">Copy prompt to clipboard</button>
        <span id="copy-feedback" class="copy-feedback"></span>
      </div>
    </div>

    <div class="card">
      <h2>Paste ChatGPT response (JSON)</h2>
      <textarea id="json" placeholder="Paste JSON array or batch object here…" rows="10"></textarea>
      <div class="btn-row">
        <button onclick="addBatch()">Add Batch</button>
        <button onclick="refreshStatus()">Refresh Status</button>
        <button onclick="exportExcel()">Export Excel</button>
        <button id="btn-visualize" type="button" disabled onclick="openVisualize()" title="Open a visual view of an exported Excel file">Visualize</button>
      </div>
    </div>

    <div id="modal-pick-file" class="modal" style="display:none;">
      <div class="modal-backdrop" onclick="closePickFileModal()"></div>
      <div class="modal-box">
        <h3>Choose Excel file to visualize</h3>
        <select id="select-excel-file" class="modal-select"></select>
        <div class="modal-actions">
          <button type="button" onclick="confirmPickFile()">Open</button>
          <button type="button" class="secondary" onclick="closePickFileModal()">Cancel</button>
        </div>
      </div>
    </div>

    <div id="status" class="row"></div>
    <pre id="preview"></pre>
  </div>

<script>
async function addBatch() {
  const txt = document.getElementById('json').value;
  const res = await fetch('/api/add-batch', { method:'POST', headers:{'Content-Type':'application/json'}, body: JSON.stringify({json: txt}) });
  const data = await res.json();
  show(data, res.ok);
}

let excelFilesList = [];

async function refreshStatus() {
  const res = await fetch('/api/status');
  const data = await res.json();
  show(data, res.ok);
  updateExcelButton();
}

async function updateExcelButton() {
  try {
    const res = await fetch('/api/excel-files');
    const data = await res.json();
    excelFilesList = data.files || [];
    const btn = document.getElementById('btn-visualize');
    btn.disabled = excelFilesList.length === 0;
  } catch (e) {
    excelFilesList = [];
    document.getElementById('btn-visualize').disabled = true;
  }
}

function openVisualize() {
  if (excelFilesList.length === 0) return;
  if (excelFilesList.length === 1) {
    window.open('/view?file=' + encodeURIComponent(excelFilesList[0].name), '_blank');
    return;
  }
  const sel = document.getElementById('select-excel-file');
  sel.innerHTML = '';
  excelFilesList.forEach(f => {
    const o = document.createElement('option');
    o.value = f.name;
    o.textContent = f.name;
    sel.appendChild(o);
  });
  document.getElementById('modal-pick-file').style.display = 'flex';
}

function closePickFileModal() {
  document.getElementById('modal-pick-file').style.display = 'none';
}

function confirmPickFile() {
  const sel = document.getElementById('select-excel-file');
  const name = sel.value;
  closePickFileModal();
  if (name) window.open('/view?file=' + encodeURIComponent(name), '_blank');
}

async function exportExcel() {
  const res = await fetch('/api/export', { method:'POST' });
  const data = await res.json();
  show(data, res.ok);
  if (data && data.output_path) {
    alert("Exported: " + data.output_path);
    updateExcelButton();
  }
}

function show(data, ok) {
  const status = document.getElementById('status');
  status.className = ok ? 'ok' : 'bad';
  status.textContent = ok ? 'OK' : 'ERROR';
  document.getElementById('preview').textContent = JSON.stringify(data, null, 2);
}

function copyChatGPTHint() {
  const ta = document.getElementById('chatgpt-hint');
  const feedback = document.getElementById('copy-feedback');
  ta.select();
  ta.setSelectionRange(0, 99999);
  navigator.clipboard.writeText(ta.value).then(function() {
    feedback.textContent = 'Copied.';
    setTimeout(function() { feedback.textContent = ''; }, 2000);
  }).catch(function() {
    feedback.textContent = 'Copy failed (try selecting and Ctrl+C).';
  });
}
updateExcelButton();
refreshStatus();
</script>
</body>
</html>
"""
        html = html.replace("___CHATGPT_HINT_ESCAPED___", html_module.escape(CHATGPT_HINT))
        return Response(html, mimetype="text/html")

    @app.get("/api/status")
    def status():
        batches = _load_session_batches()
        merged, merge_warnings = merge_batches(batches)
        return jsonify({
            "session_file": str(_session_file()),
            "batches_count": len(batches),
            "merged_items_count": len(merged),
            "merge_warnings": merge_warnings,
        })

    @app.get("/api/excel-files")
    def excel_files():
        """List .xlsx files in output dir for Visualize button."""
        files = _list_excel_files()
        return jsonify({"files": files})

    @app.get("/api/excel-data")
    def excel_data():
        """Read one Excel file's Extracted sheet; ?file=basename (must be in output dir)."""
        name = request.args.get("file", "").strip()
        if not name or ".." in name or "/" in name or "\\" in name:
            return jsonify({"error": "Invalid file name"}), 400
        path = (_OUTPUT_DIR / name).resolve()
        if not path.is_file() or path.parent != _OUTPUT_DIR.resolve():
            return jsonify({"error": "File not found"}), 404
        try:
            headers, rows, summary = _read_excel_extracted(path)
            return jsonify({"headers": headers, "rows": rows, "summary": summary})
        except Exception as e:
            return jsonify({"error": str(e)}), 400

    @app.get("/view")
    def view():
        """Visualize an Excel file; ?file=basename. Renders dashboard that fetches data via API."""
        name = request.args.get("file", "").strip()
        if not name or ".." in name or "/" in name or "\\" in name:
            return Response("<h1>Invalid file</h1>", status=400)
        path = (_OUTPUT_DIR / name).resolve()
        if not path.is_file() or path.parent != _OUTPUT_DIR.resolve():
            return Response("<h1>File not found</h1>", status=404)
        # Serve a self-contained HTML page that fetches /api/excel-data?file=... and renders it
        view_html = _view_dashboard_html(name)
        return Response(view_html, mimetype="text/html")

    @app.post("/api/add-batch")
    def add_batch():
        payload = request.get_json(force=True) or {}
        raw = payload.get("json", "")
        if not raw or not str(raw).strip():
            return jsonify({"error": "Empty JSON"}), 400

        cleaned, clean_warnings = repair_json_text(raw)

        try:
            items = json.loads(cleaned)
        except Exception as e:
            return jsonify({
                "error": f"Invalid JSON after auto-clean: {e}",
                "clean_warnings": clean_warnings,
                "cleaned_preview": cleaned[:1200],
            }), 400

        ok, errors, normalized_items = validate_batch(items)
        if not ok:
            return jsonify({"error": "Batch validation failed", "errors": errors}), 400

        _append_session_line({
            "type": "batch",
            "timestamp": datetime.now().isoformat(),
            "items": normalized_items,
        })

        batches = _load_session_batches()
        merged, merge_warnings = merge_batches(batches)

        return jsonify({
            "added": len(normalized_items),
            "batches_count": len(batches),
            "merged_items_count": len(merged),
            "merge_warnings": merge_warnings,
            "clean_warnings": clean_warnings,
        })


    @app.post("/api/export")
    def export():
        batches = _load_session_batches()
        if not batches:
            return jsonify({"error": "No batches in session yet. Add one first."}), 400

        merged, merge_warnings = merge_batches(batches)
        rows = _to_rows(merged)

        output_dir = _REPO_ROOT / "output"
        output_dir.mkdir(parents=True, exist_ok=True)
        out_path = output_dir / default_output_name(output_dir)

        warnings_list: List[str] = []
        warnings_list.extend(merge_warnings)

        export_to_excel(rows, out_path, run_warnings=warnings_list, run_failures=0)

        return jsonify({
            "output_path": str(out_path.resolve()),
            "rows": len(rows),
            "needs_review": sum(1 for r in rows if r.is_low_confidence_or_missing_key_fields()),
            "merge_warnings": merge_warnings,
        })

    url = f"http://127.0.0.1:{port}/"
    print(f"Starting OCRFactura Web UI: {url}")
    try:
        webbrowser.open(url)
    except Exception:
        pass

    # Disable Flask reloader issues on Windows PowerShell by default
    app.run(host="127.0.0.1", port=port, debug=False, use_reloader=False)
